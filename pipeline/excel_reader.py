"""
Excel 回读模块
==============
读取用户编辑后的 Excel 核对表，将修改合并回 JSON 骨架。
匹配策略：首列 Key 匹配 + 内容列位置映射。
"""

import re
import openpyxl

# Sheet name -> chapter key
SHEET_CHAPTER_MAP = {
    '第1章 客户核心画像': 'chapter1',
    '第2章 银企合作情况': 'chapter2',
    '第3章 金融需求挖掘': 'chapter3',
    '第4章 服务方案设计': 'chapter4',
    '第5章 行动计划':   'chapter5',
    '第6章 审核签发':   'chapter6',
}

SKIP_SHEETS = {'0_使用说明', '封面信息', '数据统计'}

# Known key column names
KEY_COLS = {'信息项', '分析维度', '指标', '财务指标', '动向类别', '年份',
            '债券类型', '需求类型', '维度', '产品类别', '风险类别', '准入方式',
            '合作维度', '未覆盖业务领域', '短板类别', '目标维度', '拓展方向',
            '序号', '接触类型', '角色', '银行', '职务', '子公司名称', '产品/服务',
            '联动/创新类型', '对标维度', '准入方式', '年份'}


def _strip_emoji(s: str) -> str:
    for e in ['🟢', '🟡', '🔴']:
        s = s.replace(e, '')
    return s.strip()


def _is_data_row(row_values: list) -> bool:
    """判断是否是数据行：任意列包含 🟢/🟡/🔴"""
    for v in row_values:
        if v and any(e in v for e in ('🟢', '🟡', '🔴')):
            return True
    return False


def _get_row_cols(row) -> list:
    """提取行所有列的字符串值"""
    vals = []
    for c in row:
        try:
            vals.append(str(c.value).strip() if c.value is not None else '')
        except (AttributeError, TypeError):
            vals.append('')
    return vals


def read_excel_changes(excel_path: str, data: dict, verbose: bool = False) -> int:
    """
    读取用户编辑的 Excel，将修改合并回 JSON。
    返回实际修改的行数。
    """
    wb = openpyxl.load_workbook(excel_path)
    modified = 0

    for sheet_name, ch_key in SHEET_CHAPTER_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ch_data = data['chapters'].get(ch_key, {})
        if not ch_data:
            continue

        # Collect all JSON rows into a flat search space
        # (section, table_idx, row_idx, row)
        json_rows = []
        for sec_key, sec_val in ch_data.items():
            if not isinstance(sec_val, dict):
                continue
            for ti, tbl in enumerate(sec_val.get('tables', [])):
                for ri, row in enumerate(tbl.get('data', [])):
                    json_rows.append((sec_key, ti, ri, row, tbl))

        if not json_rows:
            continue

        # Track matched JSON positions to handle duplicate keys
        key_positions = {}  # {key_val: next_index_to_match}

        # Walk all rows in sheet, match data rows to JSON
        for excel_row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            vals = _get_row_cols(excel_row)

            if not _is_data_row(vals):
                continue

            # First cell is the key (after stripping emoji if present)
            key_val = _strip_emoji(vals[0])
            if not key_val or key_val in KEY_COLS:
                continue

            # Find matching JSON row (positional for duplicate keys)
            matched = None
            pos = key_positions.get(key_val, 0)  # Which occurrence of this key
            occurrence = 0

            for sec_key, ti, ri, jrow, tbl in json_rows:
                for kc in KEY_COLS:
                    if kc in jrow and _strip_emoji(str(jrow[kc])) == key_val:
                        if occurrence == pos:
                            matched = (sec_key, ti, ri, jrow, tbl)
                            break
                        occurrence += 1
                if matched:
                    break

            if matched:
                key_positions[key_val] = pos + 1  # Next time, match next occurrence

            if not matched:
                continue

            _, _, _, jrow, _ = matched

            # Excel column structure: A=Key, B=Content, C=Source/备注, D=Status emoji
            # Only read column B (index 1) — that's the user-editable content
            if len(vals) < 2:
                continue
            excel_val = vals[1].strip()  # Column B only

            # Skip status-only values that excel_renderer wrote (not user data)
            if excel_val in ('🟢已确认', '🟡待审核', '🔴待填写', '已确认', '待审核', '待填写'):
                continue

            # Find the content column in JSON (first non-key, non-source, non-status column)
            json_col = None
            source_cols = {'备注/来源', '备注', '数据来源', '来源'}
            for k in jrow:
                if k.startswith('_'):
                    continue
                if k in KEY_COLS or k in source_cols:
                    continue
                json_col = k
                break

            if not json_col:
                continue

            json_val = str(jrow.get(json_col, '')).strip()

            if excel_val == json_val:
                continue
            if not excel_val and not json_val:
                continue

            # CHANGE DETECTED
            if verbose:
                print(f'  [{sheet_name}] {key_val}: "{json_val[:40]}" -> "{excel_val[:40]}"')

            jrow[json_col] = excel_val
            old_status = jrow.get('_status', 'red')

            if old_status == 'red' and excel_val:
                jrow['_status'] = 'green'
            elif old_status == 'green' and not excel_val:
                jrow['_status'] = 'red'

            modified += 1

    wb.close()
    return modified
