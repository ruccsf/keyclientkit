"""
JSON → Excel 核对表生成器（V2）
==============================
从标准化 JSON 生成多 Sheet Excel 核对表。
★ 这是人机协作的"落地文件" — 行内人员在此审核、修正、补充数据。

Sheet 结构:
  0. 使用说明 — 操作指引
  1. 封面信息 — 客户基本信息
  2. 第1章～第6章 — 每章一个 Sheet
  末. 数据统计 — 🟢🟡🔴 汇总

用法:
  python pipeline_v2/excel_renderer.py --client 中粮集团
  python pipeline_v2/excel_renderer.py --json output_v2/data/中粮集团_data.json
"""

import sys, os, json, re, argparse
from pathlib import Path
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

WORKSPACE = Path(__file__).parent.parent
OUTPUT_DIR = WORKSPACE / 'output_v2'

# ============================================================
# 样式
# ============================================================
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
TITLE_FONT = Font(bold=True, size=14, color='C0392B')
SECTION_FONT = Font(bold=True, size=12, color='2C3E50')

GREEN_FILL = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')
RED_FILL = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

STATUS_FILLS = {'green': GREEN_FILL, 'yellow': YELLOW_FILL, 'red': RED_FILL}

# ============================================================
# 文本清洗
# ============================================================
def clean_for_excel(val):
    """清洗单元格文本：去除 Markdown/HTML 标记，给业务人员看纯文本"""
    if val is None:
        return ''
    s = str(val)
    # HTML 实体
    for e, c in [('&amp;','&'),('&lt;','<'),('&gt;','>'),('&quot;','"'),('&nbsp;',' ')]:
        s = s.replace(e, c)
    # Markdown
    s = re.sub(r'\*\*(.+?)\*\*', r'\1', s)
    s = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', s)  # 链接保留文本丢URL
    s = re.sub(r'[🟢🟡🔴]', '', s)
    s = re.sub(r'[［\[].*?[］\]]', '', s)
    return s.strip()


# ============================================================
# Sheet 构建
# ============================================================
def write_instruction_sheet(wb: Workbook):
    """使用说明 Sheet"""
    ws = wb.active
    ws.title = '0_使用说明'
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60

    instructions = [
        ('📋 集团客户合作策略报告 — Excel 核对表', ''),
        ('', ''),
        ('🔴🟡🟢 红黄绿灯说明', ''),
        ('🟢 绿色底色', '公开数据，AI 已自动采集。请快速复核是否准确。'),
        ('🟡 黄色底色', 'AI 推断/分析数据。请审核并修正。'),
        ('🔴 红色底色', '银行内部数据。请手工填写您掌握的信息（授信/存款/合作历史等）。'),
        ('', ''),
        ('📝 操作步骤', ''),
        ('第1步', '逐 Sheet 审核数据：🟢复核、🟡修正、🔴填写。'),
        ('第2步', '修改完成后保存 Excel（Ctrl+S）。'),
        ('第3步', '运行: python pipeline_v2/run_v2.py --client <客户名> --from-excel'),
        ('第4步', '系统自动从 Excel 回写 JSON 并生成最终 HTML 报告。'),
        ('', ''),
        ('⚠️ 注意事项', ''),
        ('', '不要修改 Sheet 名称和表头行。'),
        ('', '🔴 内部数据请勿通过外网传输，仅在本地 Excel 中填写。'),
        ('', f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'),
    ]

    for i, (a, b) in enumerate(instructions, 1):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True, size=12) if i == 1 else Font(size=11)
        ws.cell(row=i, column=2, value=b)


def write_cover_sheet(wb: Workbook, data: dict):
    """封面信息 Sheet"""
    ws = wb.create_sheet('封面信息')
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15

    # 标题
    meta = data.get('meta', {})
    ws.merge_cells('A1:C1')
    ws.cell(row=1, column=1, value=f'{meta.get("client_name", "")} — 合作策略报告').font = TITLE_FONT

    # 表头
    for ci, h in enumerate(['字段', '内容', '状态'], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER

    # 封面数据行
    cover = data.get('cover', {})
    row = 4
    for key in ['客户名称', '客户等级', '所属行业', '所属分行/支行', '客户经理', '报告日期', '版本号']:
        val_info = cover.get(key, {})
        if isinstance(val_info, dict):
            value = clean_for_excel(val_info.get('value', ''))
            status = val_info.get('status', 'red')
        else:
            value = clean_for_excel(str(val_info)) if val_info else ''
            status = 'red'

        st = status if status in STATUS_FILLS else 'red'
        for ci, v in enumerate([key, value, {'green':'🟢已确认','yellow':'🟡待审核','red':'🔴待填写'}.get(st, '🔴')], 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = STATUS_FILLS.get(st, RED_FILL)
            c.border = THIN_BORDER
            c.alignment = WRAP
        row += 1

    ws.auto_filter.ref = f'A3:C{row-1}'


def write_chapter_sheet(wb: Workbook, ch_key: str, ch_data: dict, ch_name: str):
    """一个章节一个 Sheet"""
    # Sheet 名限制 31 字符
    sheet_name = ch_name[:28]
    ws = wb.create_sheet(sheet_name)

    # 列宽
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12

    row = 1

    for sec_title, sec_val in ch_data.items():
        if not isinstance(sec_val, dict):
            continue

        # 子节标题
        if sec_title != '_default':
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value=clean_for_excel(sec_title)).font = SECTION_FONT
            row += 1

        for tbl in sec_val.get('tables', []):
            tbl_title = tbl.get('title', '')
            data_rows = tbl.get('data', [])
            if not data_rows:
                continue

            # 表标题
            if tbl_title:
                ws.merge_cells(f'A{row}:D{row}')
                clean_title = clean_for_excel(re.sub(r'^[（\(]?\d+[）\)\.]?\s*', '', tbl_title))
                ws.cell(row=row, column=1, value=clean_title).font = Font(bold=True, color='C0392B')
                row += 1

            # 表头
            display_cols = [k for k in data_rows[0].keys() if not k.startswith('_')]
            for ci, h in enumerate(display_cols + ['状态'], 1):
                val = h if h != '状态' else '状态'
                c = ws.cell(row=row, column=ci, value=val)
                c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER
            row += 1

            # 数据行
            for drow in data_rows:
                st = drow.get('_status') or 'red'
                if st not in STATUS_FILLS:
                    st = 'red'
                fill = STATUS_FILLS[st]

                for ci, col_name in enumerate(display_cols, 1):
                    val = clean_for_excel(drow.get(col_name, ''))
                    c = ws.cell(row=row, column=ci, value=val)
                    c.fill = fill; c.border = THIN_BORDER; c.alignment = WRAP

                # 状态列
                status_label = {'green':'🟢已确认','yellow':'🟡待审核','red':'🔴待填写'}.get(st, '🔴')
                c = ws.cell(row=row, column=len(display_cols) + 1, value=status_label)
                c.fill = fill; c.border = THIN_BORDER; c.alignment = CENTER

                row += 1

            row += 1  # 表间空行

    # 自动筛选
    if row > 3:
        ws.auto_filter.ref = f'A1:D{row-1}'

    # 冻结表头（第一行标题之后）
    ws.freeze_panes = 'A2'


def write_stats_sheet(wb: Workbook, data: dict):
    """数据统计 Sheet"""
    ws = wb.create_sheet('数据统计')

    green = yellow = red = 0
    ch_stats = {}

    for ch_key, ch_val in data.get('chapters', {}).items():
        if not isinstance(ch_val, dict):
            continue
        g = y = r = 0
        for sec_val in ch_val.values():
            if not isinstance(sec_val, dict):
                continue
            for tbl in sec_val.get('tables', []):
                for row in tbl.get('data', []):
                    st = row.get('_status') or 'red'
                    if st == 'green': g += 1
                    elif st == 'yellow': y += 1
                    else: r += 1
        ch_stats[ch_key] = (g, y, r)
        green += g; yellow += y; red += r

    total = green + yellow + red
    if total == 0:
        ws.cell(row=1, column=1, value='暂无数据')
        return

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # 总览
    ws.merge_cells('A1:E1')
    ws.cell(row=1, column=1, value=f'数据统计 — {data.get("meta",{}).get("client_name","")}').font = TITLE_FONT

    for ci, h in enumerate(['章节', '🟢 公开数据', '🟡 AI推断', '🔴 待填写', '合计'], 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER

    ch_names = {'chapter1':'第1章 客户核心画像','chapter2':'第2章 银企合作情况',
                'chapter3':'第3章 金融需求挖掘','chapter4':'第4章 服务方案设计',
                'chapter5':'第5章 行动计划','chapter6':'第6章 审核签发'}
    row = 4
    for ch_key in ['chapter1','chapter2','chapter3','chapter4','chapter5','chapter6']:
        g, y, r = ch_stats.get(ch_key, (0, 0, 0))
        for ci, v in enumerate([ch_names.get(ch_key, ch_key), g, y, r, g+y+r], 1):
            ws.cell(row=row, column=ci, value=v).border = THIN_BORDER
        row += 1

    # 合计行
    for ci, v in enumerate(['合计', green, yellow, red, total], 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = Font(bold=True); c.border = THIN_BORDER

    # 填充率
    row += 2
    pct = (green + yellow) / total * 100 if total > 0 else 0
    ws.cell(row=row, column=1, value=f'AI 采集填充率: {pct:.0f}% ({green+yellow}/{total})').font = Font(bold=True, size=12)
    ws.cell(row=row+1, column=1, value=f'公开数据占比: {green/total*100:.0f}%' if total > 0 else '')


def _flatten_qcc(obj, prefix='', max_depth=5, max_items=500):
    """递归展平 QCC API 响应为 (路径, 值) 列表"""
    rows = []
    if max_depth <= 0:
        return [(prefix, str(obj)[:200])]

    if isinstance(obj, dict):
        for k, v in obj.items():
            if k.startswith('_'):
                continue
            path = f'{prefix}.{k}' if prefix else k
            rows.extend(_flatten_qcc(v, path, max_depth - 1, max_items - len(rows)))
            if len(rows) >= max_items:
                break
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            path = f'{prefix}[{i}]'
            rows.extend(_flatten_qcc(v, path, max_depth - 1, max_items - len(rows)))
            if len(rows) >= max_items:
                break
    else:
        val = str(obj) if obj is not None else ''
        if len(val) > 500:
            val = val[:500] + '…'
        rows.append((prefix, val))

    return rows


def write_qcc_raw_sheet(wb, data: dict):
    """企查查全部接口数据 Sheet — 原始 API 响应展平"""
    raw = data.get('_qcc_raw', [])
    if not raw:
        return

    ws = wb.create_sheet('企查查全部接口数据')

    # 列宽
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 55

    # 表头
    headers = ['接口名称', '中文备注', '字段路径', '值']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = WRAP

    row = 2
    for entry in raw:
        tool = entry.get('tool', '')
        label = entry.get('label', '')
        resp = entry.get('response')

        if resp is None:
            ws.cell(row=row, column=1, value=tool).border = THIN_BORDER
            ws.cell(row=row, column=2, value=label).border = THIN_BORDER
            ws.cell(row=row, column=3, value='(无数据)').border = THIN_BORDER
            ws.cell(row=row, column=4, value='').border = THIN_BORDER
            row += 1
            continue

        if isinstance(resp, dict) and resp.get('error'):
            ws.cell(row=row, column=1, value=tool).border = THIN_BORDER
            ws.cell(row=row, column=2, value=label).border = THIN_BORDER
            ws.cell(row=row, column=3, value='[错误]').border = THIN_BORDER
            ws.cell(row=row, column=4, value=str(resp.get('error', ''))[:500]).border = THIN_BORDER
            row += 1
            continue

        # 展平
        flat = _flatten_qcc(resp)
        if not flat:
            ws.cell(row=row, column=1, value=tool).border = THIN_BORDER
            ws.cell(row=row, column=2, value=label).border = THIN_BORDER
            ws.cell(row=row, column=3, value='(空响应)').border = THIN_BORDER
            row += 1
            continue

        for path, val in flat:
            ws.cell(row=row, column=1, value=tool).border = THIN_BORDER
            ws.cell(row=row, column=2, value=label).border = THIN_BORDER
            ws.cell(row=row, column=3, value=path).border = THIN_BORDER
            ws.cell(row=row, column=4, value=val).border = THIN_BORDER
            ws.cell(row=row, column=4).alignment = WRAP
            row += 1

        # 接口间空行
        row += 1

    # 冻结表头
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:D{row - 1}'


# ============================================================
# 主入口
# ============================================================
def generate_excel(data: dict, output_path: str) -> str:
    """从标准化 JSON 生成 Excel 核对表"""
    wb = Workbook()

    meta = data.get('meta', {})
    client_name = meta.get('client_name', '未知客户')

    # Sheet 0: 使用说明
    write_instruction_sheet(wb)

    # Sheet 1: 封面信息
    write_cover_sheet(wb, data)

    # Sheet 2-7: 6个章节
    CHAPTER_NAMES = {
        'chapter1': '第1章 客户核心画像',
        'chapter2': '第2章 银企合作情况',
        'chapter3': '第3章 金融需求挖掘',
        'chapter4': '第4章 服务方案设计',
        'chapter5': '第5章 行动计划',
        'chapter6': '第6章 审核签发',
    }

    for ch_key, ch_name in CHAPTER_NAMES.items():
        ch_data = data.get('chapters', {}).get(ch_key, {})
        if ch_data:
            write_chapter_sheet(wb, ch_key, ch_data, ch_name)

    # 企查查原始数据 Sheet
    write_qcc_raw_sheet(wb, data)

    # 末 Sheet: 数据统计
    write_stats_sheet(wb, data)

    # 保存
    wb.save(output_path)
    sz_kb = os.path.getsize(output_path) // 1024
    print(f'  Excel 核对表: {output_path} ({sz_kb}KB, {len(wb.sheetnames)} sheets)')
    print(f'  Sheets: {", ".join(wb.sheetnames)}')

    return output_path


if __name__ == '__main__':
    sys.stdout.reconfigure(encoding='utf-8')

    parser = argparse.ArgumentParser(description='JSON → Excel 核对表（落地文件）')
    parser.add_argument('--client', '-c', help='客户名称')
    parser.add_argument('--json', '-j', help='JSON 文件路径（优先于 --client）')
    args = parser.parse_args()

    if args.json:
        json_path = Path(args.json)
    elif args.client:
        json_path = OUTPUT_DIR / 'data' / f'{args.client}_data.json'
        if not json_path.exists():
            # 回退到旧 data/
            json_path = WORKSPACE / 'data' / f'{args.client}_data.json'
        if not json_path.exists():
            print(f'[错误] 未找到数据: {json_path}')
            sys.exit(1)
    else:
        print('[错误] 需要 --client 或 --json')
        sys.exit(1)

    print(f'读取 JSON: {json_path}')
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    client = data.get('meta', {}).get('client_name', args.client or '未知')
    output_path = OUTPUT_DIR / f'{client}_核对表.xlsx'
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    generate_excel(data, str(output_path))
